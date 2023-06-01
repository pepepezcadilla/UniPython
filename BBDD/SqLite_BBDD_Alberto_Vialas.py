import sqlite3
import openpyxl
import os


def crear_base_de_datos():
    # Conectar a la base de datos (se creará automáticamente si no existe)
    conn = sqlite3.connect('database.db')
    c = conn.cursor()

    # Crear la tabla Empleados
    c.execute('''CREATE TABLE IF NOT EXISTS EMPLEADOS
                 (ID INT PRIMARY KEY,
                 NOMBRE TEXT,
                 APELLIDO TEXT,
                 DNI TEXT,
                 TELEFONO TEXT,
                 DIRECCION TEXT,
                 BANCO TEXT,
                 SALARIO REAL,
                 PUESTO TEXT)''')

    # Crear la tabla Sede
    c.execute('''CREATE TABLE IF NOT EXISTS SEDE
                 (ID INT PRIMARY KEY,
                 NOMBRE TEXT,
                 PAIS TEXT,
                 CIUDAD TEXT,
                 PROVEEDOR TEXT,
                 ID_EMPLEADO INT,
                 NOMBRE_EMPLEADO TEXT,
                 APELLIDO_EMPLEADO TEXT)''')

    # Crear la tabla I+D+I
    c.execute('''CREATE TABLE IF NOT EXISTS I_D_I
                 (ID_IDI INT PRIMARY KEY,
                 ID_PROYECTO TEXT,
                 FECHAREVISION TEXT,
                 FECHAENTREGA TEXT,
                 II_IDEMPLEADOS INT,
                 II_NEMPLEADO TEXT,
                 II_AEMPLEADO TEXT)''')
    
    # Crear la tabla Aux
    c.execute('''CREATE TABLE IF NOT EXISTS Aux
                 (id_Aux INT PRIMARY KEY,
                 nombre TEXT,
                 apellido TEXT)''')

    # Guardar los cambios y cerrar la conexión
    conn.commit()
    conn.close()


def insertar_datos():
    # Conectar a la base de datos
    conn = sqlite3.connect('database.db')
    c = conn.cursor()

    # Leer los datos del archivo Excel
    workbook = openpyxl.load_workbook('BBDD SQLite-1.xlsx')
    empleados_sheet = workbook['EMPLEADOS']
    sede_sheet = workbook['SEDE']
    idi_sheet = workbook['I+D+I']

    # Insertar los datos en la tabla Empleados
    for row in empleados_sheet.iter_rows(min_row=2, values_only=True):
        if row[-1] is None:
            row = row[:-1]
        # Convertir los valores numéricos en cadenas de texto
        row = [str(value) for value in row]
        c.execute('INSERT INTO EMPLEADOS VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)', row)

    # Insertar los datos en la tabla Sede
    for row in sede_sheet.iter_rows(min_row=2, values_only=True):
        # Convertir los valores numéricos en cadenas de texto
        row = [str(value) for value in row]
        c.execute('INSERT INTO SEDE VALUES (?, ?, ?, ?, ?, ?, ?, ?)', row)

    # Insertar los datos en la tabla I+D+I
    for row in idi_sheet.iter_rows(min_row=2, values_only=True):
        if row[-1] is None and row[-2] is None:
            row = row[:-2]
        # Convertir los valores numéricos en cadenas de texto
        row = [str(value) for value in row]
        c.execute('INSERT INTO I_D_I VALUES (?, ?, ?, ?, ?, ?, ?)', row)

    # Insertar los empleados cuyo dni sea mayor de 22000000 en la tabla Aux
    c.execute('''INSERT INTO Aux (id_Aux, nombre, apellido)
                 SELECT ID, NOMBRE, APELLIDO
                 FROM Empleados
                 WHERE CAST(DNI AS INT) > 22000000''')

    # Guardar los cambios y cerrar la conexión
    conn.commit()
    conn.close()

def contar_empleados_mayor_2000():
    # Conectar a la base de datos
    conn = sqlite3.connect('database.db')
    c = conn.cursor()

    # Contar los empleados que ganan más de 2000 euros al mes
    c.execute('SELECT COUNT(*) FROM Empleados WHERE SALARIO > 2000')
    count = c.fetchone()[0]

    # Cerrar la conexión
    conn.close()

    return count


def contar_empleados_desarrollador_mayor_2000():
    # Conectar a la base de datos
    conn = sqlite3.connect('database.db')
    c = conn.cursor()

    # Contar los empleados que ganan más de 2000 euros al mes y sean "developer"
    c.execute('SELECT COUNT(*) FROM Empleados WHERE SALARIO > 2000 AND PUESTO = "DEVELOPER"')
    count = c.fetchone()[0]

    # Cerrar la conexión
    conn.close()

    return count


def obtener_dni_empleados_proyecto_1():
    # Conectar a la base de datos
    conn = sqlite3.connect('database.db')
    c = conn.cursor()

    # Obtener el dni de los empleados que trabajan en i+d+i, ganan más de 2000 euros al mes y trabajan en el proyecto 1
    c.execute('''
        SELECT CAST(DNI AS INTEGER)
        FROM EMPLEADOS
        WHERE ID IN (
            SELECT II_IDEMPLEADOS
            FROM I_D_I
            WHERE ID_PROYECTO = 'PROYECTO 1'
        ) AND SALARIO > 2000
    ''')
    dni_list = c.fetchall()
    dni_list = ', '.join(str(item[0]) for item in dni_list)

    # Cerrar la conexión
    conn.close()

    return dni_list


def obtener_dni_empleados_proyecto_2():
    # Conectar a la base de datos
    conn = sqlite3.connect('database.db')
    c = conn.cursor()

    # Obtener el dni de los empleados que trabajan en i+d+i, ganan más de 3000 euros al mes y trabajan en el proyecto 2
    c.execute('''
        SELECT CAST(DNI AS INTEGER)
        FROM EMPLEADOS
        WHERE ID IN (
            SELECT II_IDEMPLEADOS
            FROM I_D_I
            WHERE ID_PROYECTO = 'PROYECTO 2'
        ) AND SALARIO > 3000
    ''')
    dni_list = c.fetchall()
    dni_list = ', '.join(str(item[0]) for item in dni_list)

    # Cerrar la conexión
    conn.close()

    return dni_list


def obtener_bancos_empleados_mayor_4000():
    # Conectar a la base de datos
    conn = sqlite3.connect('database.db')
    c = conn.cursor()

    # Obtener la entidad bancaria de los empleados que ganan más de 4000 euros al mes, trabajan en los proyectos 1, 2 o 3 y cuyo país sea España
    c.execute('''
        SELECT E.BANCO
        FROM EMPLEADOS E, I_D_I I, SEDE S
        WHERE E.SALARIO > 4000
            AND E.ID = I.II_IDEMPLEADOS
            AND E.ID = S.ID_EMPLEADO
            AND I.ID_PROYECTO IN ('PROYECTO 1', 'PROYECTO 2', 'PROYECTO 3')
            AND UPPER(S.PAIS)= 'ESPAÑA'
    ''')

    bank_list = c.fetchall()
    bank_list = ', '.join(str(item[0]) for item in bank_list)

    # Cerrar la conexión
    conn.close()

    return bank_list


def obtener_empleados_en_ambas_tablas():
    # Conectar a la base de datos
    conn = sqlite3.connect('database.db')
    c = conn.cursor()

    # Obtener el dni de los empleados que están en la tabla Empleados y en la tabla i+d+i
    c.execute('SELECT CAST(E.DNI AS INTEGER) FROM EMPLEADOS E, I_D_I I WHERE E.ID = I.II_IDEMPLEADOS')
    dni_list = c.fetchall()
    dni_list = ', '.join(str(item[0]) for item in dni_list)

    # Cerrar la conexión
    conn.close()

    return dni_list


def obtener_empleados_no_en_tabla_empleados():
    # Conectar a la base de datos
    conn = sqlite3.connect('database.db')
    c = conn.cursor()

    # Obtener el dni de los empleados que NO están en la tabla Empleados y en la tabla i+d+i (lo he puesto con la id porque no hay DNI en la tabla i+d+i)
    c.execute("SELECT DISTINCT II_IDEMPLEADOS FROM I_D_I WHERE II_IDEMPLEADOS NOT IN (SELECT ID FROM EMPLEADOS)")

    dni_list = c.fetchall()
    dni_list = ', '.join(str(item[0]) for item in dni_list)

    # Cerrar la conexión
    conn.close()

    return dni_list


def obtener_todos_los_id_empleados():
    # Conectar a la base de datos
    conn = sqlite3.connect('database.db')
    c = conn.cursor()

    # Obtener el dni de todos los empleados que están en la tabla Empleados más los que no están en la tabla Empleados pero sí están en la tabla i+d+i (igual que arriba, cojo el ID)
    c.execute("SELECT ID FROM EMPLEADOS UNION SELECT II_IDEMPLEADOS FROM I_D_I")

    dni_list = c.fetchall()
    dni_list = ', '.join(str(item[0]) for item in dni_list)

    # Cerrar la conexión
    conn.close()

    return dni_list

def imprimir_datos_tabla(nombre):
    # Conectar a la base de datos
    conn = sqlite3.connect('database.db')
    c = conn.cursor()

    # Obtener todos los registros de la tabla
    c.execute(f'SELECT * FROM {nombre}')
    rows = c.fetchall()

    # Imprimir los registros
    for row in rows:
        print(row)

    # Cerrar la conexión
    conn.close()

# Borrar la base de datos anterior
if os.path.exists("database.db"):
        os.remove("database.db")

# Crear la base de datos
crear_base_de_datos()

# Insertar los valores del archivo Excel en la base de datos
insertar_datos()

# Ejemplo de uso de las funciones:
print("Número de empleados que ganan más de 2000 euros al mes:", contar_empleados_mayor_2000())
print("Número de empleados 'developer' que ganan más de 2000 euros al mes:", contar_empleados_desarrollador_mayor_2000())  
print("DNI de los empleados que trabajan en i+d+i, ganan más de 2000 euros al mes y trabajan en el proyecto 1:", obtener_dni_empleados_proyecto_1())
print("DNI de los empleados que trabajan en i+d+i, ganan más de 3000 euros al mes y trabajan en el proyecto 2:", obtener_dni_empleados_proyecto_2())
print("Entidad bancaria de los empleados que ganan más de 4000 euros al mes, trabajan en los proyectos 1, 2 o 3 y cuyo país es España:", obtener_bancos_empleados_mayor_4000())
print("DNI de los empleados que están en la tabla Empleados y en la tabla i+d+i:", obtener_empleados_en_ambas_tablas())
print("DNI de los empleados que NO están en la tabla Empleados y en la tabla i+d+i:", obtener_empleados_no_en_tabla_empleados())
print("DNI de todos los empleados que están en la tabla Empleados más los que no están en la tabla Empleados pero sí están en la tabla i+d+i:", obtener_todos_los_id_empleados())
